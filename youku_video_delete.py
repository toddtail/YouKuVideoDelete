import requests
from openpyxl import load_workbook

# replace the delete_list.xls with you own xls' file name
xls_data = load_workbook(filename="delete_list.xlsx")
xls_table = xls_data.worksheets[0]
xls_rows = xls_table.max_row


def run():

    url_del = 'https://openapi.youku.com/v2/videos/destroy.json'

    # replace your private parameter youku openapi client_id below, ignore the cideo_id
    payloads_del = {
        'client_id':'9**********0b',
        'access_token': '',
        'video_id':'str'
        }

    delete_videos(xls_rows,url_del,payloads_del)


def delete_videos(rows,url_del,payloads_del):
    print('Start delete')

    for i in range(1,rows+1):
        id = xls_table.cell(row=i,column=1).value
        payloads_del['video_id'] = id
        print(payloads_del)
        del_result = requests.get(url_del, payloads_del).json()
        print(del_result)

    print('finished')


run()